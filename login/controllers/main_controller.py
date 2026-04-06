import os
import subprocess
import flet as ft
from datetime import datetime
from models.user_model import UserModel
from models.permiso_model import PermisoModel
from views.admin_view import AdminView
from views.permission_view import PermissionView
from views.detail_view import DetailView
from views.login_view import LoginView
from utils.logger import logger
from utils.estado_utils import nombre_completo


class MainController:
    def __init__(self, page):
        self.page = page
        self.login_view = None
        self.usuario_actual = None
        PermisoModel.create_table()
        UserModel.create_table()

    def set_login_view(self, login_view):
        self.login_view = login_view

    def _hora_actual(self):
        return datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    def _show_error(self, mensaje):
        snack = ft.SnackBar(
            ft.Text(mensaje),
            bgcolor=ft.Colors.RED_700,
            duration=4000,
        )
        self.page.overlay.append(snack)
        snack.open = True
        self.page.update()

    def attempt_login(self, username, password):
        usuario = UserModel.authenticate(username, password)
        if usuario:
            self.usuario_actual = usuario
            logger.info(
                "INICIO DE SESION EXITOSO | Usuario: %s | Nombre completo: %s | Rol: %s | Fecha y hora: %s",
                usuario["username"],
                usuario["nombre"],
                usuario["rol"],
                self._hora_actual(),
            )
            self.show_admin_panel()
        else:
            logger.warning(
                "INTENTO DE SESION FALLIDO | Usuario intentado: %s | Contrasena incorrecta | Fecha y hora: %s",
                username,
                self._hora_actual(),
            )
            if self.login_view:
                self.login_view.show_error("Usuario o contrasena incorrectos")

    def show_admin_panel(self):
        try:
            self.page.clean()
            permisos = PermisoModel.get_all()
            nombre = self.usuario_actual["nombre"] if self.usuario_actual else "Admin"
            logger.info(
                "PANEL ABIERTO | Usuario: %s | Total de permisos cargados: %d",
                nombre,
                len(permisos),
            )
            admin_view = AdminView(
                on_add_permission=self.show_permission_panel,
                lista_permisos=permisos,
                on_edit=self.show_edit_panel,
                on_delete=self.delete_permiso,
                on_view_detail=self.show_detail_panel,
                on_logout=self.logout,
                usuario=self.usuario_actual,
                on_export=self.exportar_excel,
                on_backup=self.crear_backup,
                on_change_password=self.cambiar_password,
            )
            self.page.add(admin_view)
            self.page.update()
        except Exception as e:
            logger.error("ERROR al abrir el panel de administracion: %s", e)
            self._show_error("Error al cargar el panel: %s" % e)

    def show_permission_panel(self):
        self.page.clean()
        permission_view = PermissionView(
            on_save=self.save_permiso,
            on_back=self.show_admin_panel,
        )
        self.page.add(permission_view)
        self.page.update()

    def save_permiso(self, datos: dict):
        try:
            nuevo_id = PermisoModel.save(datos)
            nombre = nombre_completo(datos)
            usuario = self.usuario_actual["nombre"] if self.usuario_actual else "Desconocido"
            logger.info(
                "PERMISO CREADO | ID del registro: %d | Persona: %s | Tipo de permiso: %s | Periodo: del %s al %s | Registrado por: %s",
                nuevo_id,
                nombre,
                datos.get("tipo_permiso", "No especificado"),
                datos.get("fecha_desde", "No especificada"),
                datos.get("fecha_hasta", "No especificada"),
                usuario,
            )
            self.show_admin_panel()
        except Exception as e:
            logger.error("ERROR al crear el permiso: %s", e)
            self._show_error("Error al guardar el permiso: %s" % e)

    def show_edit_panel(self, permiso_id: int):
        try:
            self.page.clean()
            datos = PermisoModel.get_by_id(permiso_id)
            logger.info(
                "EDICION INICIADA | Permiso ID: %d | Persona: %s",
                permiso_id,
                nombre_completo(datos),
            )
            permission_view = PermissionView(
                on_save=lambda d: self.update_permiso(permiso_id, d),
                on_back=self.show_admin_panel,
                datos_iniciales=datos,
                permiso_id=permiso_id,
            )
            self.page.add(permission_view)
            self.page.update()
        except Exception as e:
            logger.error("ERROR al cargar el permiso para editar (ID: %d): %s", permiso_id, e)
            self._show_error("Error al cargar el permiso: %s" % e)

    def show_detail_panel(self, permiso_id: int):
        try:
            self.page.clean()
            datos = PermisoModel.get_by_id(permiso_id)
            logger.info(
                "DETALLE CONSULTADO | Permiso ID: %d | Persona: %s | Tipo: %s",
                permiso_id,
                nombre_completo(datos),
                datos.get("tipo_permiso", "No especificado"),
            )
            detail_view = DetailView(
                datos=datos,
                on_back=self.show_admin_panel,
                on_edit=self.show_edit_panel,
            )
            self.page.add(detail_view)
            self.page.update()
        except Exception as e:
            logger.error("ERROR al cargar el detalle del permiso (ID: %d): %s", permiso_id, e)
            self._show_error("Error al cargar el detalle: %s" % e)

    def update_permiso(self, permiso_id: int, datos: dict):
        try:
            PermisoModel.update(permiso_id, datos)
            usuario = self.usuario_actual["nombre"] if self.usuario_actual else "Desconocido"
            logger.info(
                "PERMISO ACTUALIZADO | ID del registro: %d | Persona: %s | Tipo de permiso: %s | Modificado por: %s",
                permiso_id,
                nombre_completo(datos),
                datos.get("tipo_permiso", "No especificado"),
                usuario,
            )
            self.show_admin_panel()
        except Exception as e:
            logger.error("ERROR al actualizar el permiso (ID: %d): %s", permiso_id, e)
            self._show_error("Error al actualizar el permiso: %s" % e)

    def delete_permiso(self, permiso_id: int):
        try:
            datos = PermisoModel.get_by_id(permiso_id)
            PermisoModel.delete(permiso_id)
            usuario = self.usuario_actual["nombre"] if self.usuario_actual else "Desconocido"
            logger.info(
                "PERMISO ELIMINADO | ID del registro: %d | Persona: %s | Eliminado por: %s",
                permiso_id,
                nombre_completo(datos),
                usuario,
            )
            self.show_admin_panel()
        except Exception as e:
            logger.error("ERROR al eliminar el permiso (ID: %d): %s", permiso_id, e)
            self._show_error("Error al eliminar el permiso: %s" % e)

    def logout(self):
        nombre = self.usuario_actual["nombre"] if self.usuario_actual else "Desconocido"
        usuario = self.usuario_actual["username"] if self.usuario_actual else "N/A"
        logger.info(
            "CIERRE DE SESION | Usuario: %s | Nombre: %s | Fecha y hora: %s",
            usuario,
            nombre,
            self._hora_actual(),
        )
        self.usuario_actual = None
        self.page.clean()
        login_view = LoginView(on_login_click=self.attempt_login)
        self.set_login_view(login_view)
        self.page.add(login_view)
        self.page.update()

    def exportar_excel(self, permisos, template_key="resumen"):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from utils.constants import EXPORT_TEMPLATES

        try:
            template = EXPORT_TEMPLATES.get(template_key)
            if not template:
                self._show_error("Template de exportacion no valido")
                return

            columnas = template["columnas"]
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Permisos"

            verde_oscuro = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            verde_claro = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            texto_blanco = Font(color="FFFFFF", bold=True, size=10)
            texto_normal = Font(size=10)
            texto_negro_bold = Font(bold=True, size=10)
            titulo_font = Font(color="2E7D32", bold=True, size=16)
            subtitulo_font = Font(color="757575", size=11)
            borde = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            num_cols = len(columnas)

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
            celda_titulo = ws.cell(row=1, column=1, value="SISTEMA DE VACACIONES - Reporte de Permisos")
            celda_titulo.font = titulo_font
            celda_titulo.alignment = Alignment(horizontal="center")

            nombre_usuario = self.usuario_actual["nombre"] if self.usuario_actual else "Admin"
            fecha_export = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
            celda_sub = ws.cell(row=2, column=1, value="Exportado: %s por: %s" % (fecha_export, nombre_usuario))
            celda_sub.font = subtitulo_font
            celda_sub.alignment = Alignment(horizontal="center")

            for col_idx, (_, titulo) in enumerate(columnas, 1):
                celda = ws.cell(row=4, column=col_idx, value=titulo)
                celda.font = texto_blanco
                celda.fill = verde_oscuro
                celda.alignment = Alignment(horizontal="center", vertical="center")
                celda.border = borde

            for fila_idx, p in enumerate(permisos, 5):
                fill = verde_claro if (fila_idx - 5) % 2 == 0 else blanco
                for col_idx, (campo, _) in enumerate(columnas, 1):
                    valor = p.get(campo, "") or ""
                    celda = ws.cell(row=fila_idx, column=col_idx, value=str(valor))
                    celda.font = texto_normal
                    celda.fill = fill
                    celda.border = borde

            total_fila = len(permisos) + 5
            ws.merge_cells(start_row=total_fila, start_column=1, end_row=total_fila, end_column=num_cols)
            celda_total = ws.cell(row=total_fila, column=1, value="Total de registros: %d" % len(permisos))
            celda_total.font = texto_negro_bold
            celda_total.fill = verde_claro
            celda_total.alignment = Alignment(horizontal="center")

            for col_idx in range(1, num_cols + 1):
                max_len = 12
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=4, max_row=total_fila - 1):
                    for cell in row:
                        if cell.value:
                            max_len = max(max_len, min(len(str(cell.value)), 35))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = "permisos_%s.xlsx" % timestamp
            export_dir = os.path.join(os.path.dirname(__file__), '..', 'data', 'export')
            os.makedirs(export_dir, exist_ok=True)
            filepath = os.path.join(export_dir, filename)
            wb.save(filepath)

            logger.info(
                "EXPORTACION REALIZADA | Template: %s | Registros: %d | Archivo: %s | Por: %s",
                template_key,
                len(permisos),
                os.path.basename(filepath),
                nombre_usuario,
            )

            snack = ft.SnackBar(
                content=ft.Row([
                    ft.Text("Exportados %d registros: %s" % (len(permisos), os.path.basename(filepath)), expand=True),
                    ft.TextButton(
                        "Abrir carpeta",
                        style=ft.ButtonStyle(
                            color=ft.Colors.WHITE,
                            bgcolor=ft.Colors.GREEN_600,
                            shape=ft.RoundedRectangleBorder(radius=6),
                        ),
                        on_click=lambda e: self._abrir_carpeta(os.path.dirname(filepath)),
                    ),
                ]),
                bgcolor=ft.Colors.GREEN_700,
                duration=6000,
            )
            self.page.overlay.append(snack)
            snack.open = True
            self.page.update()

        except ImportError:
            self._show_error("Error: openpyxl no esta instalado. Ejecute: pip install openpyxl")
        except Exception as e:
            logger.error("ERROR al exportar a Excel: %s", e)
            self._show_error("Error al exportar: %s" % e)

    def _abrir_carpeta(self, ruta):
        try:
            if os.name == "nt":
                os.startfile(ruta)
            elif os.name == "posix":
                subprocess.Popen(["open" if os.uname().sysname == "Darwin" else "xdg-open", ruta])
        except (OSError, FileNotFoundError) as e:
            self._show_error("No se pudo abrir la carpeta: %s" % e)

    def crear_backup(self):
        from utils.backup import crear_backup
        try:
            nombre = crear_backup()
            logger.info("BACKUP CREADO | Nombre: %s | Por: %s", nombre, self.usuario_actual.get("nombre", "Admin") if self.usuario_actual else "Desconocido")
            snack = ft.SnackBar(
                content=ft.Row([
                    ft.Text("Backup creado: %s" % nombre, expand=True),
                    ft.TextButton(
                        "Abrir carpeta",
                        style=ft.ButtonStyle(
                            color=ft.Colors.WHITE,
                            bgcolor=ft.Colors.GREEN_600,
                            shape=ft.RoundedRectangleBorder(radius=6),
                        ),
                        on_click=lambda e: self._abrir_carpeta(os.path.join(os.path.dirname(__file__), '..', 'data', 'backups')),
                    ),
                ]),
                bgcolor=ft.Colors.GREEN_700,
                duration=6000,
            )
            self.page.overlay.append(snack)
            snack.open = True
            self.page.update()
        except Exception as e:
            logger.error("ERROR al crear backup: %s", e)
            self._show_error("Error al crear backup: %s" % e)

    def cambiar_password(self):
        from models.user_model import UserModel, verify_password

        input_actual = ft.TextField(
            label="Contrasena actual",
            password=True,
            can_reveal_password=True,
        )
        input_nueva = ft.TextField(
            label="Nueva contrasena",
            password=True,
            can_reveal_password=True,
        )
        input_confirmar = ft.TextField(
            label="Confirmar nueva contrasena",
            password=True,
            can_reveal_password=True,
        )

        def cerrar_dialogo(e):
            dialogo.open = False
            self.page.update()

        def actualizar(e):
            if not input_actual.value:
                self._show_error("Ingrese la contrasena actual")
                return
            if not input_nueva.value:
                self._show_error("Ingrese la nueva contrasena")
                return
            if input_nueva.value != input_confirmar.value:
                self._show_error("Las contrasenas nuevas no coinciden")
                return
            if len(input_nueva.value) < 4:
                self._show_error("La nueva contrasena debe tener al menos 4 caracteres")
                return

            usuario = self.usuario_actual
            row = UserModel.authenticate(usuario["username"], input_actual.value)
            if not row:
                self._show_error("La contrasena actual es incorrecta")
                return

            UserModel.cambiar_password(usuario["id"], input_nueva.value)
            logger.info("CONTRASENA CAMBIADA | Usuario: %s | Por: %s", usuario["username"], usuario["nombre"])
            dialogo.open = False
            self.page.update()
            snack = ft.SnackBar(
                ft.Text("Contrasena actualizada correctamente"),
                bgcolor=ft.Colors.GREEN_700,
                duration=4000,
            )
            self.page.overlay.append(snack)
            snack.open = True
            self.page.update()

        dialogo = ft.AlertDialog(
            title=ft.Row([
                ft.Icon(ft.Icons.LOCK_OUTLINE, color=ft.Colors.GREEN_700, size=24),
                ft.Text("Cambiar Contrasena", size=18, weight=ft.FontWeight.BOLD),
            ], spacing=10),
            content=ft.Column([
                input_actual,
                input_nueva,
                input_confirmar,
            ], tight=True, spacing=12),
            actions=[
                ft.TextButton("Cancelar", on_click=cerrar_dialogo),
                ft.ElevatedButton(
                    "Actualizar",
                    icon=ft.Icons.CHECK_CIRCLE,
                    style=ft.ButtonStyle(
                        color=ft.Colors.WHITE,
                        bgcolor=ft.Colors.GREEN_700,
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                    on_click=actualizar,
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )

        self.page.overlay.append(dialogo)
        dialogo.open = True
        self.page.update()
